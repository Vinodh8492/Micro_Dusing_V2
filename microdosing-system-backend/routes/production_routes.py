from flask import Blueprint, request, jsonify # type: ignore
from extensions import db
from models.production import ProductionOrder, Batch, BatchMaterialDispensing
from flask_jwt_extended import jwt_required, get_jwt_identity # type: ignore
from routes.user_routes import role_required  # Adjust path based on your project structure
from flask import Blueprint, send_file, jsonify
from openpyxl import Workbook
from openpyxl.drawing.image import Image as ExcelImage
from barcode import Code128
from barcode.writer import ImageWriter
from PIL import Image as PILImage
import os, io, tempfile


production_bp = Blueprint("production", __name__)

@production_bp.route("/production_orders/export/barcodes", methods=["GET"])
def export_production_orders_excel_with_barcodes():
    try:
        orders = ProductionOrder.query.all()

        wb = Workbook()
        ws = wb.active
        ws.title = "Production Order Barcodes"
        ws.append(["Order Number", "Barcode ID", "Scannable Barcode"])

        row_number = 2

        for order in orders:
            if order.barcode_id:
                barcode_id = order.barcode_id
                try:
                    # Save barcode image to temp dir
                    temp_dir = tempfile.gettempdir()
                    filename = f"{barcode_id}"
                    filepath = os.path.join(temp_dir, f"{filename}.png")

                    code128 = Code128(barcode_id, writer=ImageWriter())
                    code128.save(filepath)

                    # Resize image
                    image = PILImage.open(filepath)
                    image = image.resize((200, 60))
                    image.save(filepath)

                    # Add to Excel
                    ws.cell(row=row_number, column=1, value=order.order_number)
                    ws.cell(row=row_number, column=2, value=barcode_id)

                    img = ExcelImage(filepath)
                    img.width = 150
                    img.height = 50
                    ws.add_image(img, f"C{row_number}")

                    # Clean up temp file
                    os.remove(filepath)

                    row_number += 1

                except Exception as e:
                    print(f"Failed to generate barcode for {barcode_id}: {e}")
                    ws.cell(row=row_number, column=1, value=order.order_number)
                    ws.cell(row=row_number, column=2, value=barcode_id)
                    row_number += 1

        stream = io.BytesIO()
        wb.save(stream)
        stream.seek(0)

        return send_file(
            stream,
            download_name="production_orders_with_barcodes.xlsx",
            as_attachment=True,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({"error": str(e)}), 500



@production_bp.route("/production_orders", methods=["POST"])
@jwt_required(locations=["headers"])
@role_required(["admin", "operator"])  # Only allowed roles can create
def create_production_order():
    data = request.get_json()

    required_fields = ["order_number", "recipe_id", "batch_size", "scheduled_date"]
    missing = [field for field in required_fields if field not in data]
    if missing:
        return jsonify({"error": f"Missing fields: {', '.join(missing)}"}), 400

    current_user_id = get_jwt_identity()

    try:
        new_order = ProductionOrder(
            order_number=data["order_number"],
            recipe_id=data["recipe_id"],
            batch_size=data["batch_size"],
            scheduled_date=data["scheduled_date"],
            status="planned",
            created_by=current_user_id,
            notes=data.get("notes"),
            barcode_id=data.get("barcode_id"),
        )
        db.session.add(new_order)
        db.session.commit()
        return jsonify({"message": "Production order created successfully!"}), 201
    except Exception as e:
        db.session.rollback()
        if "Duplicate entry" in str(e):
            return jsonify({"error": "Duplicate order/barcode"}), 400
        return jsonify({"error": "Failed to create order", "details": str(e)}), 500

@production_bp.route("/production_orders/<int:order_id>", methods=["PUT"])
def update_production_order(order_id):
    data = request.get_json()

    # Debugging Step: Print received data
    print("Received Data for Update:", data)

    if not data:
        return jsonify({"error": "No data received"}), 400

    # Fetch the existing order by ID
    order = ProductionOrder.query.get(order_id)

    if not order:
        return jsonify({"error": "Production order not found"}), 404

    try:
        # Update order fields with new data
        order.order_number = data.get("order_number", order.order_number)
        order.recipe_id = data.get("recipe_id", order.recipe_id)
        order.batch_size = data.get("batch_size", order.batch_size)
        order.scheduled_date = data.get("scheduled_date", order.scheduled_date)
        order.status = data.get("status", order.status)
        order.created_by = data.get("created_by", order.created_by)
        order.notes = data.get("notes", order.notes)

        db.session.commit()
        return jsonify({"message": "Production order updated successfully!"}), 200

    except Exception as e:
        db.session.rollback()
        print("Error updating data:", str(e))
        return jsonify({"error": str(e)}), 500

@production_bp.route("/production_orders/<int:order_id>", methods=["DELETE"])
def delete_production_order(order_id):
    try:
        # Find the production order by ID
        order = ProductionOrder.query.get(order_id)
        
        if not order:
            return jsonify({"error": "Production order not found"}), 404

        # Delete the order
        db.session.delete(order)
        db.session.commit()

        return jsonify({"message": f"Production order {order_id} deleted successfully!"}), 200

    except Exception as e:
        db.session.rollback()
        print("Error deleting data:", str(e))
        return jsonify({"error": str(e)}), 500


@production_bp.route("/production_orders", methods=["GET"])
def get_production_orders():
    orders = ProductionOrder.query.all()
    result = [
        {
            "order_id": order.order_id,
            "order_number": order.order_number,
            "recipe_id": order.recipe_id,
            "batch_size": str(order.batch_size),
            "scheduled_date": order.scheduled_date.strftime("%Y-%m-%d"),
            "status": order.status,
            "created_by": order.created_by,
        }
        for order in orders
    ]
    return jsonify(result)

@production_bp.route("/production-orders/<int:order_id>/reject", methods=["PUT"])
@jwt_required()
@role_required(["admin"])  # ✅ Only admin can reject
def reject_production_order(order_id):
    order = ProductionOrder.query.get(order_id)

    if not order:
        return jsonify({"error": "Production order not found"}), 404

    order.status = "rejected"
    db.session.commit()

    return jsonify({"message": "Production order rejected successfully"}), 200

### 🚀 BATCH ROUTES ###
@production_bp.route("/batches", methods=["POST"])
def create_batch():
    data = request.get_json()
    new_batch = Batch(
        batch_number=data["batch_number"],
        order_id=data["order_id"],
        status=data.get("status", "pending"),
        operator_id=data["operator_id"],
        notes=data.get("notes"),
    )
    db.session.add(new_batch)
    db.session.commit()
    return jsonify({"message": "Batch created successfully!"}), 201

@production_bp.route("/batches", methods=["GET"])
def get_batches():
    batches = Batch.query.all()
    result = [
        {
            "batch_id": batch.batch_id,
            "batch_number": batch.batch_number,
            "order_id": batch.order_id,
            "status": batch.status,
            "operator_id": batch.operator_id,
            "notes": batch.notes,
            "created_at": batch.created_at,  # Add created_by field here
        }
        for batch in batches
    ]
    return jsonify(result)

### 🚀 BATCH MATERIAL DISPENSING ROUTES ###
@production_bp.route("/batch_dispensing", methods=["POST"])
def create_batch_dispensing():
    data = request.get_json()
    new_dispensing = BatchMaterialDispensing(
        batch_id=data["batch_id"],
        material_id=data["material_id"],
        planned_quantity=data["planned_quantity"],
        actual_quantity=data.get("actual_quantity"),
        dispensed_by=data["dispensed_by"],
        status=data.get("status", "pending"),
    )
    db.session.add(new_dispensing)
    db.session.commit()
    return jsonify({"message": "Material dispensing record created successfully!"}), 201

@production_bp.route("/batch_dispensing", methods=["GET"])
def get_batch_dispensing():
    dispensing_records = BatchMaterialDispensing.query.all()
    result = [
        {
            "dispensing_id": record.dispensing_id,
            "batch_id": record.batch_id,
            "material_id": record.material_id,
            "planned_quantity": str(record.planned_quantity),
            "actual_quantity": str(record.actual_quantity) if record.actual_quantity else None,
            "dispensed_by": record.dispensed_by,
            "status": record.status,
        }
        for record in dispensing_records
    ]
    return jsonify(result)
